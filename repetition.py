import openpyxl
from openpyxl.styles import Font, colors, PatternFill
import os


def read_disscussed(file_path):
    wb0 = openpyxl.load_workbook(file_path)
    ws0 = wb0["Sheet1"]
    dic_bold = {}
    bold_labels = []
    blue_labels = []
    for row in range(1, ws0.max_row):
        if ws0.cell(row, 2).font.bold:
            bold_labels.append(ws0.cell(row, 2).value)
            dic_bold[ws0.cell(row, 2).value] = {column: ws0.cell(row, column).value for column in range(1, ws0.max_column+1)}
        try:
            if ws0.cell(row, 2).font.color.rgb == 'FF0070C0':
                blue_labels.append(ws0.cell(row, 2).value)
        except:
            pass
    
    return dic_bold, bold_labels, blue_labels

def write_repetition(file_path, dic_bold, bold_labels, blue_labels):
    list_column = [1, 3, 4]
    font_bold = Font(
        size=12.0,
        italic=None,
        color=None,
        bold=True,
        strike=None
    )
    font_blue = Font(
        size=12.0,
        italic=None,
        color='FF0070C0',
        bold=None,
        strike=None
    )
    font_bold_blue = Font(
        size=12.0,
        italic=None,
        color='FF0070C0',
        bold=True,
        strike=None
    )
    
    wb1 = openpyxl.load_workbook(file_path)
    ws1 = wb1["Sheet1"]
    
    for row in range(1, ws1.max_row):
        if ws1.cell(row, 2).value in bold_labels and ws1.cell(row, 2).value in blue_labels:
            ws1.cell(row, 2).font = font_bold_blue
            for column in list_column:
                ws1.cell(row, column).value = dic_bold[ws1.cell(row, 2).value][column]
        elif ws1.cell(row, 2).value in bold_labels:
            ws1.cell(row, 2).font = font_bold
            for column in range(1, ws1.max_column):
                ws1.cell(row, column).value = dic_bold[ws1.cell(row, 2).value][column]
        elif ws1.cell(row, 2).value in blue_labels:
            ws1.cell(row, 2).font = font_blue
    
    wb1.save(file_path)
    

dic_bold, bold_labels, blue_labels = read_disscussed("./汇总.xlsx")
projects = ['req', 'devops', 'acp', 'ait', 'asm', 'bjlyq', 'middle']
for project in projects:
    write_repetition("./data_each_project/%s.xlsx"%project, dic_bold, bold_labels, blue_labels)